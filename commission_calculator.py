import openpyxl
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Prompt the user to select the XLSX file
root = Tk()
root.withdraw()
xlsx_file = askopenfilename(title="Select the XLSX file", filetypes=[("Excel files", "*.xlsx")])

# Load the Excel file
workbook = openpyxl.load_workbook(xlsx_file)
sheet = workbook.active

# Create a dictionary to store the totals for each category
category_totals = {}

# Iterate through the rows in the sheet
for row in range(2, sheet.max_row + 1):
    order_date = sheet.cell(row=row, column=1).value
    category = sheet.cell(row=row, column=2).value
    item_name = sheet.cell(row=row, column=3).value
    quantity = sheet.cell(row=row, column=4).value
    item_cost = sheet.cell(row=row, column=5).value
    btcusd = sheet.cell(row=row, column=6).value

    # Check if the value in the "Item Cost" column is a number
    if not isinstance(item_cost, (int, float)):
        try:
            item_cost = float(item_cost)
        except ValueError:
            print(f"Warning: Non-numeric value '{item_cost}' found in 'Item Cost' column for row {row}. Skipping this row.")
            continue

    # Check if the value in the "btcusd" column is None
    if btcusd is None:
        print(f"Warning: Empty value found in 'btcusd' column for row {row}. Skipping this row.")
        continue

    # Calculate the sats per dollar
    sats_per_dollar = 1 / (btcusd / 100_000_000)

    # Calculate the total in sats for this sale
    total_sats = quantity * item_cost * sats_per_dollar

    # Update the category total
    if category in category_totals:
        category_totals[category] += total_sats
    else:
        category_totals[category] = total_sats

# Prompt the user to enter the commission percentage for each category
commission_percentages = {}
for category in category_totals:
    commission_percentage = float(input(f"Enter the commission percentage for category '{category}': "))
    commission_percentages[category] = commission_percentage

# Determine the output file name
output_file = os.path.splitext(os.path.basename(xlsx_file))[0] + "_commission_report.txt"

# Calculate the commission for each category and write to a text file
with open(output_file, 'w') as file:
    for category, total in category_totals.items():
        commission = total * (commission_percentages[category] / 100)
        file.write(f"Category: {category}\nCommission: {int(commission)} sats\n\n")

print(f"Commission report saved to: {output_file}")
